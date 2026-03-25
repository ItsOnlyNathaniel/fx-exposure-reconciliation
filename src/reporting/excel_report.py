from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Colour palette
NAVY       = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
GREEN      = "1E8449"
AMBER      = "D4AC0D"
RED        = "C0392B"
LIGHT_GREY = "F2F2F2"
DARK_GREY  = "595959"

def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_cell(ws, row, col, value, bg=NAVY, fg=WHITE, size=11, bold=True, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True
    )
    cell.border = _border()
    return cell

def _data_cell(ws, row, col, value, bold=False, color=None, bg=WHITE, fmt=None, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, color=color or "000000", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center"
    )
    cell.border = _border()
    if fmt:
        cell.number_format = fmt
    return cell

def _section_title(ws, row, col, value, span_end_col):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=MID_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _border()
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=span_end_col
    )
    return cell

def build_summary_tab(wb: Workbook, run_metadata: dict, kpis: dict) -> None:
    """
    run_metadata: {
        run_id, run_timestamp, status,
        feed_source, ledger_source, operator
    }
    kpis: {
        total_feed_trades, total_ledger_trades, matched_trades,
        total_breaks, rate_mismatches, notional_mismatches,
        settlement_mismatches, status_mismatches,
        missing_trades, duplicates,
        margin_breaches, open_positions, total_exposure_usd
    }
    """
    ws = wb.active
    ws.title = "Summary"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20

    # ── Banner ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    banner = ws["A1"]
    banner.value = "FX Reconciliation Report — Daily Summary"
    banner.font = Font(name="Arial", bold=True, color=WHITE, size=14)
    banner.fill = PatternFill("solid", start_color=NAVY)
    banner.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M:%S')}"
    sub.font = Font(name="Arial", color=DARK_GREY, size=9, italic=True)
    sub.fill = PatternFill("solid", start_color=LIGHT_GREY)
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # ── Run Metadata ─────────────────────────────────────────────────────
    _section_title(ws, row=4, col=1, value="RUN METADATA", span_end_col=4)

    meta_rows = [
        ("Run ID",          run_metadata.get("run_id", "-")),
        ("Run Timestamp",   run_metadata.get("run_timestamp", "-")),
        ("Status",          run_metadata.get("status", "-")),
        ("Feed Source",     run_metadata.get("feed_source", "-")),
        ("Ledger Source",   run_metadata.get("ledger_source", "-")),
        ("Operator",        run_metadata.get("operator", "-")),
    ]

    status_colours = {"COMPLETED": GREEN, "RUNNING": AMBER, "FAILED": RED, "PENDING": DARK_GREY}

    for i, (label, value) in enumerate(meta_rows, start=5):
        ws.row_dimensions[i].height = 18
        _data_cell(ws, i, 1, label, bold=True, bg=LIGHT_BLUE)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        if label == "Status":
            colour = status_colours.get(str(value).upper(), DARK_GREY)
            _data_cell(ws, i, 2, value, bold=True, color=colour)
        else:
            _data_cell(ws, i, 2, value)

    # ── Trade Volume KPIs ─────────────────────────────────────────────────
    _section_title(ws, row=12, col=1, value="TRADE VOLUME", span_end_col=4)
    _header_cell(ws, 13, 1, "Metric",      bg=MID_BLUE)
    _header_cell(ws, 13, 2, "Bank Feed",   bg=MID_BLUE)
    _header_cell(ws, 13, 3, "Int. Ledger", bg=MID_BLUE)
    _header_cell(ws, 13, 4, "Matched",     bg=MID_BLUE)

    volume_rows = [
        (
            "Trade Count",
            kpis.get("total_feed_trades", 0),
            kpis.get("total_ledger_trades", 0),
            kpis.get("matched_trades", 0),
        )
    ]
    for i, (label, feed, ledger, matched) in enumerate(volume_rows, start=14):
        ws.row_dimensions[i].height = 18
        _data_cell(ws, i, 1, label, bold=True, bg=LIGHT_GREY)
        _data_cell(ws, i, 2, feed,    fmt="#,##0", center=True)
        _data_cell(ws, i, 3, ledger,  fmt="#,##0", center=True)
        _data_cell(ws, i, 4, matched, fmt="#,##0", center=True, bold=True, color=GREEN)

    # ── Break Summary ─────────────────────────────────────────────────────
    _section_title(ws, row=16, col=1, value="BREAK SUMMARY", span_end_col=4)
    _header_cell(ws, 17, 1, "Break Type",  bg=MID_BLUE)
    _header_cell(ws, 17, 2, "Count",       bg=MID_BLUE)
    _header_cell(ws, 17, 3, "% of Matched",bg=MID_BLUE)
    _header_cell(ws, 17, 4, "Severity",    bg=MID_BLUE)

    matched = kpis.get("matched_trades", 1) or 1
    total_breaks = kpis.get("total_breaks", 0)

    break_rows = [
        ("Rate Mismatch",        kpis.get("rate_mismatches", 0),        "MEDIUM"),
        ("Notional Mismatch",    kpis.get("notional_mismatches", 0),    "MEDIUM"),
        ("Settlement Mismatch",  kpis.get("settlement_mismatches", 0),  "LOW"),
        ("Status Mismatch",      kpis.get("status_mismatches", 0),      "LOW"),
        ("Missing Trade",        kpis.get("missing_trades", 0),         "HIGH"),
        ("Duplicate",            kpis.get("duplicates", 0),             "HIGH"),
    ]

    severity_colours = {"HIGH": RED, "MEDIUM": AMBER, "LOW": GREEN}

    for i, (label, count, severity) in enumerate(break_rows, start=18):
        ws.row_dimensions[i].height = 18
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        _data_cell(ws, i, 1, label,  bg=bg)
        _data_cell(ws, i, 2, count,  bg=bg, fmt="#,##0", center=True, bold=(count > 0))
        pct_formula = f"=B{i}/B{i - (i - 18) + 18 + len(break_rows) + 2}" # calculated below
        _data_cell(ws, i, 3, count / matched, bg=bg, fmt="0.0%", center=True)
        sev_colour = severity_colours.get(severity, DARK_GREY)
        _data_cell(ws, i, 4, severity, bg=bg, bold=True, color=sev_colour, center=True)

    # Total breaks row
    total_row = 18 + len(break_rows)
    ws.row_dimensions[total_row].height = 20
    _data_cell(ws, total_row, 1, "TOTAL BREAKS", bold=True, bg=LIGHT_BLUE)
    _data_cell(ws, total_row, 2, f"=SUM(B18:B{total_row - 1})", bold=True, bg=LIGHT_BLUE, fmt="#,##0", center=True)
    _data_cell(ws, total_row, 3, f"=B{total_row}/{matched}", bold=True, bg=LIGHT_BLUE, fmt="0.0%", center=True)
    _data_cell(ws, total_row, 4, "", bg=LIGHT_BLUE)

    # ── Margin & Exposure ────────────────────────────────────────────────
    exp_start = total_row + 2
    _section_title(ws, row=exp_start, col=1, value="MARGIN & EXPOSURE", span_end_col=4)

    exp_rows = [
        ("Margin Breaches",      kpis.get("margin_breaches", 0),     "#,##0",        margin_colour(kpis.get("margin_breaches", 0))),
        ("Open Positions",       kpis.get("open_positions", 0),      "#,##0",        None),
        ("Total Exposure (USD)", kpis.get("total_exposure_usd", 0.0),"$#,##0.00",    None),
    ]

    for i, (label, value, fmt, colour) in enumerate(exp_rows, start=exp_start + 1):
        ws.row_dimensions[i].height = 18
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        _data_cell(ws, i, 1, label, bold=True, bg=LIGHT_BLUE)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        _data_cell(ws, i, 2, value, fmt=fmt, bold=(colour is not None), color=colour or "000000", center=True)

    # ── Freeze panes & tab colour ────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = MID_BLUE.replace("#", "")


def margin_colour(breach_count: int) -> str:
    if breach_count == 0:
        return GREEN
    elif breach_count <= 2:
        return AMBER
    return RED


def generate_report(run_metadata: dict, kpis: dict, output_path: str) -> None:
    wb = Workbook()
    build_summary_tab(wb, run_metadata, kpis)
    wb.save(output_path)


# ── Demo entrypoint ──────────────────────────────────────────────────────
if __name__ == "__main__":
    sample_metadata = {
        "run_id":         "RUN-20260323-001",
        "run_timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status":         "COMPLETED",
        "feed_source":    "bank_feed_2026-03-23.csv",
        "ledger_source":  "internal_ledger_2026-03-23.json",
        "operator":       "automated",
    }
    sample_kpis = {
        "total_feed_trades":    200,
        "total_ledger_trades":  197,
        "matched_trades":       193,
        "total_breaks":         21,
        "rate_mismatches":       6,
        "notional_mismatches":   9,
        "settlement_mismatches": 3,
        "status_mismatches":     1,
        "missing_trades":        3,
        "duplicates":            2,
        "margin_breaches":       1,
        "open_positions":       84,
        "total_exposure_usd":   4_823_150.75,
    }

    out = "recon_report_demo.xlsx"
    generate_report(sample_metadata, sample_kpis, out)
    print(f"Report written → {out}")
